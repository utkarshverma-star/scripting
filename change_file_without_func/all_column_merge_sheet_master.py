import pandas as pd
import requests
from urllib.parse import urlparse
from openpyxl import load_workbook
from dotenv import load_dotenv
import os

load_dotenv() 

SPREADSHEET_ID = os.getenv("SPREADSHEET_ID")

MASTER_DIR = os.getenv("MASTER_PATH_RECKITT")
FILENAME = "Reckitt_Creator_Master.xlsx"
FILENAME1 = "Reckitt_Creator_Master.xlsx"
FILENAME2 = "downloaded_sheet.xlsx"

MASTER_FILE = os.path.join(MASTER_DIR, FILENAME)
OUTPUT_FILE = os.path.join(MASTER_DIR, FILENAME1)
DOWNLOAD_FILE = os.path.join(MASTER_DIR, FILENAME2)

# DOWNLOAD GOOGLE SHEET

print("Downloading Google Sheet...")

download_url = (
    f"https://docs.google.com/spreadsheets/d/"
    f"{SPREADSHEET_ID}/export?format=xlsx"
)

response = requests.get(download_url, timeout=60)
response.raise_for_status()

with open(DOWNLOAD_FILE, "wb") as f:
    f.write(response.content)

print("Download completed.")

# SHORTCODE EXTRACTION
def extract_shortcode(url):
    if pd.isna(url):
        return ""

    try:
        parts = [x for x in urlparse(str(url)).path.split("/") if x]

        if len(parts) >= 2 and parts[0] in ["p", "reel", "reels"]:
            return parts[1]

    except Exception:
        pass

    return ""


# READ MASTER FILE

print("Reading master file...")

master_df = pd.read_excel(MASTER_FILE)

# Remove Media ID if exists
if "Media ID" in master_df.columns:
    master_df.drop(columns=["Media ID"], inplace=True)
# READ ALL SHEETS

print("Reading all sheets...")

excel_file = pd.ExcelFile(DOWNLOAD_FILE)

all_dfs = []

for sheet in excel_file.sheet_names:

    df = pd.read_excel(
        DOWNLOAD_FILE,
        sheet_name=sheet,
        dtype=str
    )

    print(f"{sheet}: {len(df)} rows")

    if not df.empty:
        all_dfs.append(df)

if not all_dfs:
    raise Exception("No data found in any sheet.")

source_df = pd.concat(
    all_dfs,
    ignore_index=True
)

print(f"\nTotal Source Rows: {len(source_df)}")

# CLEAN COLUMN NAMES

source_df.columns = source_df.columns.str.strip()

# RENAME COLUMNS
source_df.rename(
    columns={
        "Creator Username": "Name",
        "Live Links": "Live Link",
        "Go Live Date": "Live Date",
        "Content Bucket": "Bucket"
    },
    inplace=True
)

# LIVE DATE


if "Live Date" in source_df.columns:
    source_df["Live Date"] = pd.to_datetime(
        source_df["Live Date"],
        errors="coerce"
    ).dt.date

# SHORTCODE

if "Live Link" in source_df.columns:
    source_df["ShortCode"] = source_df["Live Link"].apply(
        extract_shortcode
    )
else:
    source_df["ShortCode"] = ""

# REMOVE MEDIA ID IF EXISTS

if "Media ID" in source_df.columns:
    source_df.drop(columns=["Media ID"], inplace=True)

# REQUIRED COLUMNS
all_columns = list(
    dict.fromkeys(
        list(master_df.columns)
        + list(source_df.columns)
    )
)

# Remove Media ID
all_columns = [
    col for col in all_columns
    if col != "Media ID"
]

master_df = master_df.reindex(columns=all_columns)
source_df = source_df.reindex(columns=all_columns)
# APPEND DATA
final_df = pd.concat(
    [master_df, source_df],
    ignore_index=True
)
# REMOVE DUPLICATES USING SHORTCODE
final_df["ShortCode"] = (
    final_df["ShortCode"]
    .fillna("")
    .astype(str)
    .str.strip()
)

rows_before_dedup = len(final_df)

has_shortcode = final_df["ShortCode"] != ""

dedup_df = final_df[has_shortcode].drop_duplicates(
    subset=["ShortCode"],
    keep="last"
)

blank_shortcode_df = final_df[~has_shortcode]

final_df = pd.concat(
    [dedup_df, blank_shortcode_df],
    ignore_index=True
)

rows_after_dedup = len(final_df)

duplicates_removed = (
    rows_before_dedup - rows_after_dedup
)

# SAVE OUTPUT

final_df.to_excel(
    OUTPUT_FILE,
    index=False
)
# FORMAT EXCEL


wb = load_workbook(OUTPUT_FILE)
ws = wb.active

headers = [cell.value for cell in ws[1]]

# Make Live Link Clickable


# if "Live Link" in headers:

#     live_link_col = headers.index("Live Link") + 1

#     for row in range(2, ws.max_row + 1):

#         cell = ws.cell(row=row, column=live_link_col)

#         if cell.value and str(cell.value).startswith("http"):
#             cell.hyperlink = str(cell.value)
#             cell.style = "Hyperlink"

# Format Live Date


if "Live Date" in headers:

    live_date_col = headers.index("Live Date") + 1

    for row in range(2, ws.max_row + 1):

        ws.cell(
            row=row,
            column=live_date_col
        ).number_format = "MM-DD-YYYY"

wb.save(OUTPUT_FILE)

# SUMMARY

print("\n========== SUMMARY ==========")
print(f"Master Rows        : {len(master_df)}")
print(f"Source Rows        : {len(source_df)}")
print(f"Rows Before Dedup  : {rows_before_dedup}")
print(f"Rows After Dedup   : {rows_after_dedup}")
print(f"Duplicates Removed : {duplicates_removed}")
print(f"\nOutput Saved : {OUTPUT_FILE}")
print("\nDone Successfully.")



