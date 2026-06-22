import pandas as pd
import requests
from bs4 import BeautifulSoup
from urllib.parse import urljoin
from requests.utils import requote_uri
from openpyxl import load_workbook
from pathlib import Path
import os
import re
import time

# =====================================================
# CONFIG
# =====================================================

ROOT_FOLDER = r"C:\change_file_without_func\file_rregulatoryh"

# =====================================================
# SESSION
# =====================================================

session = requests.Session()

session.headers.update({
    "User-Agent":
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"
})

# =====================================================
# GLOBAL DUPLICATE TRACKER
# =====================================================

downloaded_urls = set()

failed_file = os.path.join(
    ROOT_FOLDER,
    "FAILED_DOWNLOADS.txt"
)

# =====================================================
# FIND EXCEL FILES
# =====================================================

excel_files = list(Path(ROOT_FOLDER).rglob("*.xlsx"))
excel_files.extend(list(Path(ROOT_FOLDER).rglob("*.xls")))

print(f"\nExcel Files Found: {len(excel_files)}")

# =====================================================
# PROCESS FILES
# =====================================================

for excel_file in excel_files:

    print("\n" + "=" * 100)
    print("Processing:", excel_file.name)

    try:
        df = pd.read_excel(excel_file)

        wb = load_workbook(
            excel_file,
            data_only=False
        )

        ws = wb.active

    except Exception as e:
        print("Cannot read:", e)
        continue

    # =================================================
    # FIND PDF COLUMN
    # =================================================

    pdf_col = None

    for col in df.columns:

        c = str(col).strip().lower()

        if c in [
            "pdf_link",
            "pdf_url",
            "pdflink",
            "pdfurl"
        ]:
            pdf_col = col
            break

    if pdf_col is None:
        print("PDF column not found")
        continue

    print("PDF Column:", pdf_col)

    pdf_col_index = (
        list(df.columns)
        .index(pdf_col)
        + 1
    )

    # =================================================
    # TITLE COLUMN
    # =================================================

    title_col = None

    for col in df.columns:

        if (
            str(col)
            .strip()
            .lower()
            == "title"
        ):
            title_col = col
            break

    # =================================================
    # PDF FOLDER
    # =================================================

    pdf_folder = os.path.join(
        ROOT_FOLDER,
        excel_file.stem
    )

    os.makedirs(
        pdf_folder,
        exist_ok=True
    )

    success = 0
    duplicate = 0
    failed = 0

    # =================================================
    # ROW LOOP
    # =================================================

    for idx, row in df.iterrows():

        excel_row = idx + 2

        cell = ws.cell(
            row=excel_row,
            column=pdf_col_index
        )

        url = ""

        # ---------------------------------------------
        # Hyperlink handling
        # ---------------------------------------------

        if cell.hyperlink:

            url = str(
                cell.hyperlink.target
            ).strip()

        else:

            value = cell.value

            if value:
                url = str(value).strip()

        if not url:
            continue

        # ---------------------------------------------
        # Skip invalid values
        # ---------------------------------------------

        if (
            url.lower()
            in [
                "open pdf",
                "download pdf",
                "view pdf"
            ]
            and not cell.hyperlink
        ):
            continue

        # ---------------------------------------------
        # Convert relative CDSCO links
        # ---------------------------------------------

        if (
            not url.startswith("http")
            and "download_file_division.jsp"
            in url
        ):

            url = urljoin(
                "https://cdsco.gov.in",
                url
            )

        # ---------------------------------------------
        # Duplicate URL
        # ---------------------------------------------

        if url in downloaded_urls:

            duplicate += 1
            continue

        downloaded_urls.add(url)

        # ---------------------------------------------
        # Filename
        # ---------------------------------------------

        if (
            title_col
            and pd.notna(
                row[title_col]
            )
        ):
            filename = str(
                row[title_col]
            )
        else:
            filename = f"PDF_{idx+1}"

        filename = re.sub(
            r'[\\/*?:"<>|]',
            "_",
            filename
        )

        filename = filename[:180]

        file_path = os.path.join(
            pdf_folder,
            filename + ".pdf"
        )

        if os.path.exists(file_path):
            continue

        downloaded = False

        # ---------------------------------------------
        # Retry
        # ---------------------------------------------

        for attempt in range(5):

            try:

                actual_pdf_url = url

                # =====================================
                # CDSCO iframe handling
                # =====================================

                if (
                    "download_file_division.jsp"
                    in actual_pdf_url
                ):

                    r = session.get(
                        actual_pdf_url,
                        timeout=60
                    )

                    soup = BeautifulSoup(
                        r.text,
                        "html.parser"
                    )

                    iframe = soup.find(
                        "iframe"
                    )

                    if iframe:

                        actual_pdf_url = urljoin(
                            "https://cdsco.gov.in",
                            iframe.get("src")
                        )

                actual_pdf_url = requote_uri(
                    actual_pdf_url
                )

                response = session.get(
                    actual_pdf_url,
                    stream=True,
                    timeout=120
                )

                response.raise_for_status()

                content_type = (
                    response.headers.get(
                        "Content-Type",
                        ""
                    )
                )

                if (
                    "pdf"
                    not in content_type.lower()
                ):
                    raise Exception(
                        f"Not PDF -> {content_type}"
                    )

                with open(
                    file_path,
                    "wb"
                ) as f:

                    for chunk in response.iter_content(
                        8192
                    ):
                        if chunk:
                            f.write(chunk)

                size_kb = round(
                    os.path.getsize(
                        file_path
                    ) / 1024,
                    2
                )

                print(
                    f"{idx+1} Downloaded -> "
                    f"{size_kb} KB"
                )

                success += 1
                downloaded = True

                break

            except Exception as e:

                print(
                    f"{idx+1} Retry "
                    f"{attempt+1}/5 -> {e}"
                )

                time.sleep(5)

        # ---------------------------------------------
        # Failed
        # ---------------------------------------------

        if not downloaded:

            failed += 1

            with open(
                failed_file,
                "a",
                encoding="utf-8"
            ) as f:

                f.write(
                    f"{excel_file.name}"
                    f"|Row={idx+1}"
                    f"|{url}\n"
                )

        time.sleep(1)

    print("\nSummary")
    print("Downloaded :", success)
    print("Duplicate  :", duplicate)
    print("Failed     :", failed)

print("\nAll Excel Files Completed")